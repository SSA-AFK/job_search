from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook, load_workbook

from app.rankings.selection import (
    candidate_stratum,
    read_ranking_candidates,
    select_representative_sample,
)

HEADERS = [
    "公司名称",
    "登记状态",
    "企业规模",
    "成立日期",
    "所属省份",
    "所属城市",
    "国标行业大类",
    "统一社会信用代码",
    "网址",
    "天眼评分",
    "参保人数",
    "参保人数所属年报",
    "经营范围",
    "注册资本",
    "实缴资本",
    "所属区县",
    "企业(机构)类型",
    "国标行业门类",
    "国标行业中类",
]


def _workbook(path: Path, count: int = 120) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "高级搜索"
    sheet.append(["声明"])
    sheet.append(HEADERS)
    for number in range(count):
        sheet.append(
            [
                f"公司{number}",
                "存续" if number % 9 else "在业",
                "小型",
                "2020-01-01",
                f"省{number % 4}",
                f"市{number % 3}",
                f"行业{number % 3}",
                f"91310000{number:010d}",
                "-",
                str(50 + number % 40),
                str(20 + number),
                "2025",
                "人工智能软件开发",
                "1000万人民币",
                "500万人民币",
                "海淀区",
                "有限责任公司",
                "信息传输、软件和信息技术服务业",
                "软件开发",
            ]
        )
    workbook.save(path)
    return path


def _corrupt_dimension(path: Path) -> None:
    with ZipFile(path) as source:
        entries = {name: source.read(name) for name in source.namelist()}
    entries["xl/worksheets/sheet1.xml"] = entries["xl/worksheets/sheet1.xml"].replace(
        b'<dimension ref="A1:J122"/>', b'<dimension ref="A1"/>'
    )
    with ZipFile(path, "w", ZIP_DEFLATED) as target:
        for name, content in entries.items():
            target.writestr(name, content)


def test_reader_handles_bad_declared_dimension_and_discards_sensitive_columns(
    tmp_path: Path,
) -> None:
    workbook = _workbook(tmp_path / "companies.xlsx")
    _corrupt_dimension(workbook)

    candidates = read_ranking_candidates(workbook)

    assert len(candidates) == 120
    assert candidates[0].source_row == 3
    assert not hasattr(candidates[0], "phone")
    assert len(candidates[0].identity_hash) == 64
    assert candidates[0].company_size == "小型"
    assert candidates[0].established_at is not None
    assert candidates[0].insured_employee_count == 20
    assert candidates[0].employee_report_year == 2025
    assert candidates[0].business_scope == "人工智能软件开发"
    assert candidates[0].registered_capital == "1000万人民币"
    assert candidates[0].district == "海淀区"
    assert candidates[0].company_type == "有限责任公司"


def test_reader_normalizes_bare_website_domains(tmp_path: Path) -> None:
    workbook = _workbook(tmp_path / "companies.xlsx", count=1)
    loaded = load_workbook(workbook)
    loaded["高级搜索"].cell(3, HEADERS.index("网址") + 1).value = "example.com"
    loaded.save(workbook)

    assert read_ranking_candidates(workbook)[0].website_candidate == "https://example.com"


def test_hamilton_sample_is_deterministic_and_seeded(tmp_path: Path) -> None:
    candidates = read_ranking_candidates(_workbook(tmp_path / "companies.xlsx"))

    first = select_representative_sample(candidates, sample_size=20, seed="one")
    repeated = select_representative_sample(candidates, sample_size=20, seed="one")
    second = select_representative_sample(candidates, sample_size=20, seed="two")

    assert first == repeated
    assert len(first) == 20
    assert {item.source_row for item in first} != {item.source_row for item in second}
    all_strata = {candidate_stratum(item, candidates) for item in candidates}
    for stratum in all_strata:
        population = sum(candidate_stratum(item, candidates) == stratum for item in candidates)
        sampled = sum(candidate_stratum(item, candidates) == stratum for item in first)
        assert sampled in {
            20 * population // len(candidates),
            (20 * population + len(candidates) - 1) // len(candidates),
        }
