from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from app.core.normalization import normalize_name

WORKSHEET_NAME = "高级搜索"
COHORT_ROWS = range(3, 23)


class CohortWorkbookError(ValueError):
    pass


@dataclass(frozen=True)
class StagedCompany:
    canonical_name: str
    normalized_name: str
    source_row: int


def read_tianyancha_cohort(workbook_path: Path) -> tuple[StagedCompany, ...]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise CohortWorkbookError(f"worksheet {WORKSHEET_NAME} is required")
        worksheet = workbook[WORKSHEET_NAME]
        worksheet.reset_dimensions()
        result: list[StagedCompany] = []
        for source_row in COHORT_ROWS:
            value = worksheet.cell(source_row, 1).value
            if not isinstance(value, str) or not value.strip():
                raise CohortWorkbookError("rows 3-22 must contain twenty company names")
            name = value.strip()
            normalized_name = normalize_name(name)
            if not normalized_name:
                raise CohortWorkbookError("rows 3-22 must contain twenty company names")
            result.append(StagedCompany(name, normalized_name, source_row))
        return tuple(result)
    finally:
        workbook.close()
