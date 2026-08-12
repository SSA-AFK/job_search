"""Parse restricted local exports and choose a deterministic internal pilot."""

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from urllib.parse import urlsplit

from openpyxl import load_workbook

from app.core.normalization import normalize_name

WORKSHEET_NAME = "高级搜索"
_REQUIRED_HEADERS = (
    "公司名称",
    "登记状态",
    "企业规模",
    "成立日期",
    "所属省份",
    "所属城市",
    "国标行业大类",
    "统一社会信用代码",
    "网址",
    "天眼评分",
    "参保人数",
    "参保人数所属年报",
    "经营范围",
    "注册资本",
    "实缴资本",
    "所属区县",
    "企业(机构)类型",
    "国标行业门类",
    "国标行业中类",
)
_ELIGIBLE_STATUSES = frozenset({"存续", "在业"})


class RankingWorkbookError(ValueError):
    pass


@dataclass(frozen=True)
class RankingCandidate:
    canonical_name: str
    normalized_name: str
    source_row: int
    province: str
    city: str
    industry_major: str
    score: int
    identity_hash: str
    website_candidate: str | None
    company_size: str | None
    established_at: date | None
    insured_employee_count: int | None
    employee_report_year: int | None
    business_scope: str | None
    registered_capital: str | None
    paid_in_capital: str | None
    district: str | None
    company_type: str | None
    industry_sector: str | None
    industry_middle: str | None


def read_ranking_candidates(workbook_path: Path) -> tuple[RankingCandidate, ...]:
    """Read the malformed-dimension export without persisting restricted raw fields."""
    workbook = load_workbook(workbook_path, read_only=False, data_only=True)
    try:
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise RankingWorkbookError(f"worksheet {WORKSHEET_NAME} is required")
        sheet = workbook[WORKSHEET_NAME]
        headers = [sheet.cell(2, column).value for column in range(1, sheet.max_column + 1)]
        positions = _header_positions(headers)
        candidates: list[RankingCandidate] = []
        seen_names: set[str] = set()
        for row in range(3, sheet.max_row + 1):
            values = {name: sheet.cell(row, positions[name]).value for name in _REQUIRED_HEADERS}
            if _text(values["登记状态"]) not in _ELIGIBLE_STATUSES:
                continue
            name = _text(values["公司名称"])
            normalized_name = normalize_name(name)
            credit_code = _text(values["统一社会信用代码"])
            if not normalized_name or not credit_code or normalized_name in seen_names:
                continue
            score = _score(values["天眼评分"])
            if score is None:
                continue
            seen_names.add(normalized_name)
            candidates.append(
                RankingCandidate(
                    canonical_name=name,
                    normalized_name=normalized_name,
                    source_row=row,
                    province=_text(values["所属省份"]) or "未知",
                    city=_text(values["所属城市"]) or "未知",
                    industry_major=_text(values["国标行业大类"]) or "未知",
                    score=score,
                    identity_hash=sha256(credit_code.encode()).hexdigest(),
                    website_candidate=_website(values["网址"]),
                    company_size=_optional_text(values["企业规模"]),
                    established_at=_date(values["成立日期"]),
                    insured_employee_count=_integer(values["参保人数"]),
                    employee_report_year=_integer(values["参保人数所属年报"]),
                    business_scope=_optional_text(values["经营范围"]),
                    registered_capital=_optional_text(values["注册资本"]),
                    paid_in_capital=_optional_text(values["实缴资本"]),
                    district=_optional_text(values["所属区县"]),
                    company_type=_optional_text(values["企业(机构)类型"]),
                    industry_sector=_optional_text(values["国标行业门类"]),
                    industry_middle=_optional_text(values["国标行业中类"]),
                )
            )
        if not candidates:
            raise RankingWorkbookError("no eligible companies found")
        return tuple(candidates)
    finally:
        workbook.close()


def select_representative_sample(
    candidates: tuple[RankingCandidate, ...], *, sample_size: int, seed: str
) -> tuple[RankingCandidate, ...]:
    if sample_size < 1 or sample_size > len(candidates):
        raise ValueError("sample_size must be within the eligible candidate count")
    score_bucket = _score_buckets(candidates)
    groups: dict[tuple[str, str, int], list[RankingCandidate]] = {}
    for candidate in candidates:
        groups.setdefault(
            (candidate.province, candidate.industry_major, score_bucket[candidate]), []
        ).append(candidate)
    total = len(candidates)
    allocations = {key: len(group) * sample_size // total for key, group in groups.items()}
    remaining = sample_size - sum(allocations.values())
    ranked_remainders = sorted(
        groups,
        key=lambda key: (-(len(groups[key]) * sample_size % total), key),
    )
    for key in ranked_remainders[:remaining]:
        allocations[key] += 1
    selected: list[RankingCandidate] = []
    for key, group in groups.items():
        selected.extend(
            sorted(
                group, key=lambda item: sha256(f"{item.identity_hash}:{seed}".encode()).hexdigest()
            )[: allocations[key]]
        )
    return tuple(sorted(selected, key=lambda item: item.source_row))


def candidate_stratum(candidate: RankingCandidate, candidates: tuple[RankingCandidate, ...]) -> str:
    bucket = _score_buckets(candidates)[candidate]
    return f"{candidate.province}|{candidate.industry_major}|score_q{bucket + 1}"


def _header_positions(headers: list[object]) -> dict[str, int]:
    positions = {_text(value): index + 1 for index, value in enumerate(headers)}
    missing = [header for header in _REQUIRED_HEADERS if header not in positions]
    if missing:
        raise RankingWorkbookError(f"missing required header: {missing[0]}")
    return positions


def _score_buckets(candidates: tuple[RankingCandidate, ...]) -> dict[RankingCandidate, int]:
    ordered = sorted(candidates, key=lambda item: (item.score, item.identity_hash))
    return {candidate: index * 5 // len(ordered) for index, candidate in enumerate(ordered)}


def _score(value: object) -> int | None:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def _text(value: object) -> str:
    return value.strip() if isinstance(value, str) else ""


def _optional_text(value: object) -> str | None:
    text = _text(value)
    return None if text in {"", "-"} else text


def _integer(value: object) -> int | None:
    text = _optional_text(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _optional_text(value)
    if text is None:
        return None
    try:
        return date.fromisoformat(text[:10])
    except ValueError:
        return None


def _website(value: object) -> str | None:
    website = _text(value).split(";")[0].strip()
    if not website or website == "-":
        return None
    candidate = website if "://" in website else f"https://{website}"
    parsed = urlsplit(candidate)
    return candidate if parsed.scheme in {"http", "https"} and parsed.hostname else None
